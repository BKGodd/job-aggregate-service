"""
Created by: Brandon Goddard
Description: This module contains functions that relate to
             processing and loading information into the
             Elasticsearch database.
"""
from shutil import copyfileobj
from string import punctuation
from os import environ
from os.path import isfile
from uuid import uuid4
from requests.exceptions import ConnectTimeout
import requests
from openpyxl import load_workbook
from unidecode import unidecode as ud

from elasticsearch import helpers

from constants import STATES, PAY_SCALES


def build_query(inputs):
    """
    Create a custom-formatted query dictionary to be used for the
    Elasticsearch database, based on the query texts provided.

    Args:
        inputs (tuple): The query texts provided by the user
                        from the GET request, being 'title'
                        and 'location'.

    Returns:
        (dict): The ingestible query request needed for
                performing a search on Elasticsearch.
        (dict): The ingestible aggreagation request needed for
                performing aggregation calculations
                on Elasticsearch.
    """
    # Create a starting point for building the query and aggs requests
    query = {"bool": {"must": []}}
    aggs = {"salary_mean": {"avg": {"field": "salary"}},
                            "salary_percentiles":
                            {"percentiles":
                            {"field": "salary","percents": [25, 50, 75]}}}

    # Iterate through both "title" and "location" query texts
    for ind, phrase in enumerate(inputs):
        if not phrase:
            continue
        # Filter out punctuation and replace accent strings
        phrase = simplify_text(phrase)
        if ind==0:
            field = "job_title"
        else:
            field = "city_state"
        # Add each separate word into the MATCH query
        match_query = {"bool": {"should": [],
                        "minimum_should_match": "100%"}}
        for word in phrase.split():
            # We choose to NOT use wildcards here, performance not ideal
            # This should be implemented with N-gram tokenizer, not
            # implemented here due to time.
            match_query['bool']['should'].append({"match": {field: word}})
            # Commented example of a query using wildcards:
            #match_query['bool']['should'].append({"wildcard": {
            #    field: {"value": f'*{word}*'}}})
            
        # Enforce that both "title" and "location" queries are satisfied
        # for a valid hit
        query['bool']['must'].append(match_query)

    return query, aggs


def simplify_text(text):
    """
    Process the given text to be either queried or inserted into
    the Elasticsearch database.

    Args:
        text (str): The text to process and simplify for use.

    Returns:
        (str): The text after being processed.

    Notes:
        The following changes are made to the string:
            - Punctuation removed
            - Character accents removed
            - All extraneous whitespace removed
            - Lowered capitalization
    """
    if isinstance(text, str):
        text = ud(text.translate(str.maketrans('', '', punctuation)))
        return ' '.join(text.split()).lower()
    return ''


def is_valid_text(text):
    """
    Determine if given text can be inserted into database.

    Args:
        text (str): The text to check if valid.

    Returns:
        (bool): Whether the text is valid or not.
    """
    not_empty = len(text) > 0
    is_alpha = not text.isdigit()
    return not_empty and is_alpha


def filter_text(source, data, job_title_ind, job_city_ind, job_state_ind):
    """
    Gather the job title and location texts from the raw excel data.

    Args:
        source (dict): Source data that will be inserted as a document
                       into the Elasticsearch database.
        data (list): The excel row data (each element is a cell).

    Returns:
        (bool): Whether the data found within this row is satisfactory
                for insertion into Elasticsearch (not inserted if False).

    Note:
        The following columns are pulled from excel, with the corresponding
        column number (indexed from 1):
            JOB_TITLE (7)
            WORKSITE_CITY_1 (50)
            WORKSITE_STATE_1 (52)
    """
    # Look for the job title, must be found for a row to be valid
    source['job_title'] = simplify_text(data[job_title_ind])
    if not is_valid_text(source['job_title']):
        return False

    # Look for the city and state where the job resides
    city = simplify_text(data[job_city_ind])
    state = simplify_text(data[job_state_ind])
    # Convert state to non-abbreviated form, if applicable
    state = STATES.get(state.upper(), state)
    source['city_state'] = f'{city} {state}'
    # Ensure either city, state, or both are found
    if not is_valid_text(source['city_state']):
        return False

    return True


def filter_salary(source, data, wage_base_ind, wage_unit_ind):
    """
    Gather the salary from the raw excel data.

    Args:
        source (dict): Source data that will be inserted as a document
                       into the Elasticsearch database.
        data (list): The excel row data (each element is a cell).

    Returns:
        (bool): Whether the data found within this row is satisfactory
                for insertion into Elasticsearch (not inserted if False).

    Note:
        The following columns are pulled from excel, with the corresponding
        column number (indexed from 1):
            WAGE_RATE_OF_PAY_FROM_1 (54)
            WAGE_UNIT_OF_PAY_1 (56)
    """
    # Look for the base salary, must be found for a row to be valid
    if isinstance(data[wage_base_ind], (int, float)):
        salary = float(data[wage_base_ind])
    elif isinstance(data[wage_base_ind], str):
        try:
            salary = float(data[wage_base_ind])
        except ValueError as error:
            print(error)
            return False
    else:
        return False
    # Look for the unit of pay, must be found for a row to be valid
    if not isinstance(data[wage_unit_ind], str):
        return False
    pay_type = data[wage_unit_ind].strip().lower()
    # Convert the found unit of pay to the yearly equivalent
    if not PAY_SCALES.get(pay_type, None):
        return False

    # We assume that exorbitantly high salaries are mistaken, fix to year
    if pay_type != "year" and salary > 10000/PAY_SCALES[pay_type]:
        pay_type = "year"

    # The yearly equivalent is then scaled down by 1e-3, stored
    # as HALF_FLOAT in elasticsearch (max value of 65504)
    salary *= PAY_SCALES[pay_type]

    # Hard check to block salaries that can't be stored as HALF_FLOAT
    if salary > 65504000/PAY_SCALES[pay_type]:
        print(f'Salary larger than HALF_FLOAT: {salary} ({pay_type})')
        return False

    source['salary'] = salary

    return True


def get_row(excel_sheet):
    """
    Given an excel sheet instance, process all rows for
    insertion into the Elasticsearch database.

    Args:
        excel_sheet (Worksheet): Object for accessing excel file data.

    Returns:
        None
    """
    source = {}
    job_title_ind = None
    job_city_ind = None
    job_state_ind = None
    wage_base_ind = None
    wage_unit_ind = None
    for row_cells in excel_sheet.iter_rows(max_col=56):
        # Get column indices of interest from the headers
        if job_title_ind is None:
            header = [cell.value for cell in row_cells]
            job_title_ind = header.index('JOB_TITLE')
            job_city_ind = header.index('WORKSITE_CITY_1')
            job_state_ind = header.index('WORKSITE_STATE_1')
            wage_base_ind = header.index('WAGE_RATE_OF_PAY_FROM_1')
            wage_unit_ind = header.index('WAGE_UNIT_OF_PAY_1')
            continue

        # Only insert a row if sufficient query data is found
        row = [cell.value for cell in row_cells]
        if not (filter_text(source, row, job_title_ind, job_city_ind, job_state_ind) and 
                filter_salary(source, row, wage_base_ind, wage_unit_ind)):
            continue

        # Document to be loaded into Elasticsearch
        yield {
            "_id": uuid4(),
            "_index": environ['ESDB_INDEX'],
            "_source": source
        }


async def load_bulk(excel_filepath, elastic_db):
    """
    Perform a bulk load of excel data into the Elasticsearch database.

    Args:
        excel_filepath (str): Absolute path to the excel file to read.
        elastic_db (AsyncElasticsearch): Database connection instance.

    Returns:
        None
    """
    # If the file exists, load it into the Elasticsearch database
    if isfile(excel_filepath):
        excel_sheet = load_workbook(excel_filepath, read_only=True).active
        await helpers.async_bulk(elastic_db, get_row(excel_sheet))


def download_excel(excel_url, filepath):
    """
    Download an excel file as a stream.

    Args:
        excel_url: URL from which to download the excel file.
        filepath: Absolute path where the downloaded file
                  should be stored.

    Returns:
        None
    """
    try:
        with requests.get(excel_url, stream=True, timeout=500) as response:
            with open(filepath, 'wb') as local_file:
                copyfileobj(response.raw, local_file)
    except ConnectTimeout:
        print('Request to download excel file timed out.')
